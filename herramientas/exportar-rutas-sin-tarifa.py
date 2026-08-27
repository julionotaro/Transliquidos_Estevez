#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Exporta a Excel las rutas que la empresa transporta y el tarifario NO cubre.

Es la lista de trabajo para dar de alta tarifas en Gesruta, ordenada por lo que
mas rinde: primero las rutas con mas viajes al año. Dar de alta las 20 primeras
cubre mas viajes que dar de alta las 300 ultimas.

Entrada:  catalogo/rutas-por-cliente.json  (lo genera construir-matriz-rutas.py)
Salida:   informes/rutas-sin-tarifa.xlsx

Tres hojas:
  1. "Sin tarifa"      una fila por combinacion cliente x ruta x material
  2. "Resumen cliente" cuanto pesa cada cliente en el agujero
  3. "Precios inestables"  rutas que se facturaron a mas de un precio en el año
                           (el historico no puede decidir solo cual aplica)
"""
import json, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENTRADA = os.path.join(RAIZ, 'catalogo/rutas-por-cliente.json')
SALIDA = os.path.join(RAIZ, 'informes/rutas-sin-tarifa.xlsx')

CAB = Font(bold=True, color='FFFFFF')
FONDO = PatternFill('solid', fgColor='2F5496')
DESTACA = PatternFill('solid', fgColor='FFF2CC')


def encabezar(ws, cols):
    ws.append(cols)
    for i, _ in enumerate(cols, 1):
        c = ws.cell(row=1, column=i)
        c.font = CAB; c.fill = FONDO
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def anchos(ws, medidas):
    for i, w in enumerate(medidas, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    d = json.load(open(ENTRADA, encoding='utf-8'))
    filas, inestables = [], []
    por_cliente = defaultdict(lambda: [0, 0, 0])   # combinaciones, viajes, con tarifa

    for cid, c in d['clientes'].items():
        for R in c['rutas']:
            por_cliente[c['nombre']][2] += 1
            if R['en_tarifario_oficial']:
                continue
            por_cliente[c['nombre']][0] += 1
            por_cliente[c['nombre']][1] += R['n_viajes']
            filas.append([
                c['nombre'], cid, c.get('nif', ''),
                R['nombre_origen'], R['nombre_destino'], R['nombre_material'],
                R['n_viajes'], R['primera_fecha'], R['ultima_fecha'],
                R['precio_ultimo'], R['unidad'],
                len(R['precios_vistos']),
                'SI' if R['precio_estable'] else 'NO',
                R.get('causa_sin_tarifa') or '',
            ])
            if not R['precio_estable']:
                inestables.append([
                    c['nombre'], R['nombre_origen'], R['nombre_destino'], R['nombre_material'],
                    R['n_viajes'], R['unidad'],
                    ' / '.join(str(p) for p in R['precios_vistos']),
                    R['precio_ultimo'], R['ultima_fecha'],
                ])

    filas.sort(key=lambda f: -f[6])
    inestables.sort(key=lambda f: -f[4])

    wb = Workbook()

    ws = wb.active; ws.title = 'Sin tarifa'
    encabezar(ws, ['Cliente', 'Cod.cliente', 'NIF', 'Origen', 'Destino', 'Material',
                   'Viajes', 'Primer viaje', 'Ultimo viaje', 'Ultimo precio', 'U.M.',
                   'Nº precios distintos', 'Precio estable', 'Por que no hay tarifa'])
    for f in filas:
        ws.append(f)
        if f[6] >= 10:                      # las que mas rinden, resaltadas
            for i in range(1, len(f) + 1):
                ws.cell(row=ws.max_row, column=i).fill = DESTACA
    anchos(ws, [34, 11, 13, 20, 26, 16, 8, 12, 12, 13, 7, 10, 10, 46])

    ws2 = wb.create_sheet('Resumen cliente')
    encabezar(ws2, ['Cliente', 'Combinaciones sin tarifa', 'Viajes sin tarifa',
                    'Combinaciones totales', '% sin tarifa'])
    for nom, (nc, nv, tot) in sorted(por_cliente.items(), key=lambda x: -x[1][1]):
        if nc:
            ws2.append([nom, nc, nv, tot, round(100.0 * nc / tot, 1)])
    anchos(ws2, [40, 20, 16, 18, 12])

    ws3 = wb.create_sheet('Precios inestables')
    ws3.append(['Rutas que se facturaron a MAS DE UN precio durante el año. El '
                'historico no puede elegir solo: o hubo una actualizacion de tarifa, '
                'o se esta cobrando distinto por lo mismo.'])
    ws3.cell(row=1, column=1).font = Font(italic=True)
    ws3.append([])
    fila_cab = ws3.max_row + 1
    ws3.append(['Cliente', 'Origen', 'Destino', 'Material', 'Viajes', 'U.M.',
                'Precios vistos', 'Ultimo', 'Fecha ultimo'])
    for i in range(1, 10):
        c = ws3.cell(row=fila_cab, column=i); c.font = CAB; c.fill = FONDO
    for f in inestables:
        ws3.append(f)
    ws3.freeze_panes = 'A%d' % (fila_cab + 1)
    anchos(ws3, [34, 20, 26, 16, 8, 7, 40, 11, 13])

    os.makedirs(os.path.dirname(SALIDA), exist_ok=True)
    wb.save(SALIDA)
    print('escrito:', SALIDA)
    print('  sin tarifa: %d combinaciones, %d viajes' % (len(filas), sum(f[6] for f in filas)))
    print('  precios inestables: %d rutas' % len(inestables))


main()
